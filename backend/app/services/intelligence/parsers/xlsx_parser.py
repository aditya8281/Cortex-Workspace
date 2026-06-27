"""Excel spreadsheet (.xlsx) parser."""

from __future__ import annotations

import logging

from backend.app.services.intelligence.parsers.base import BaseParser, ParsedDocument, ParsedSection

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Extract data from Excel .xlsx files using openpyxl."""

    def parse(self, file_path: str) -> ParsedDocument:
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError:
            logger.warning("openpyxl not installed — run: pip install openpyxl")
            return ParsedDocument(metadata={"error": "openpyxl not installed"})

        sections: list[ParsedSection] = []
        metadata: dict = {}

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            metadata["sheet_count"] = len(workbook.sheetnames)
            metadata["sheet_names"] = workbook.sheetnames

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                # Filter out completely empty rows
                non_empty_rows = []
                for row in rows:
                    if any(cell is not None and str(cell).strip() for cell in row):
                        non_empty_rows.append(row)

                if not non_empty_rows:
                    continue

                # Build markdown table
                table_text = self._format_sheet(non_empty_rows)
                if table_text.strip():
                    sections.append(
                        ParsedSection(
                            title=sheet_name,
                            content=table_text,
                            section_type="table",
                        )
                    )

            workbook.close()

        except Exception as e:
            logger.warning("Failed to parse XLSX %s: %s", file_path, e)
            return ParsedDocument(metadata={"error": str(e)})

        total_chars = sum(len(s.content) for s in sections)
        return ParsedDocument(sections=sections, metadata=metadata, total_chars=total_chars)

    @staticmethod
    def _format_sheet(rows: list[tuple]) -> str:
        if not rows:
            return ""

        # Use first row as headers
        headers = [str(cell if cell is not None else "") for cell in rows[0]]
        data_rows = rows[1:]

        header_line = " | ".join(headers)
        separator = " | ".join(["---"] * len(headers))
        lines = [header_line, separator]

        for row in data_rows:
            cells = [str(cell if cell is not None else "") for cell in row]
            lines.append(" | ".join(cells))

        return "\n".join(lines)
